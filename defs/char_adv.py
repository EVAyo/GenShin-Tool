import os

from openpyxl import load_workbook

FILE_PATH = "assets"
char_adv_im = '''【{}】
【五星武器】：{}
【四星武器】：{}
【三星武器】：{}
【圣遗物】：
{}'''


async def weapon_adv(name):
    char_adv_path = os.path.join(FILE_PATH, "Genshin_All_Char.xlsx")
    wb = load_workbook(char_adv_path)
    ws = wb.active

    weapon_name = ""
    weapons = {}
    for c in range(2, 5):
        for r in range(2, 300):
            if ws.cell(r, c).value:
                # if all(i in ws.cell(r,c).value for i in name):
                if name in ws.cell(r, c).value:
                    weapon_name = ws.cell(r, c).value
                    weapon = weapons.get(weapon_name, [])
                    weapon.append(ws.cell(2 + ((r - 2) // 5) * 5, 1).value)
                    weapons[weapon_name] = weapon

    if weapons:
        im = []
        for k, v in weapons.items():
            im.append(f'{"、".join(v)}可能会用到【{k}】')
        im = '\n'.join(im)
    else:
        im = " 没有角色能使用【{}】".format(weapon_name if weapon_name else name)
    return im


async def char_adv(name):
    char_adv_path = os.path.join(FILE_PATH, "Genshin_All_Char.xlsx")
    wb = load_workbook(char_adv_path)
    ws = wb.active
    char_list = ws["A"]
    index = None
    for i in char_list:
        if i.value:
            if all(g in i.value for g in name):
                # if name in i.value:
                index = i.row
                char_name = i.value
    if index:
        weapon_5star = ""
        for i in range(index, index + 5):
            if ws.cell(i, 2).value:
                weapon_5star += ws.cell(i, 2).value + ">"
        if weapon_5star != "":
            weapon_5star = weapon_5star[:-1]
        else:
            weapon_5star = "无推荐"

        weapon_4star = ""
        for i in range(index, index + 5):
            if ws.cell(i, 3).value:
                weapon_4star += ws.cell(i, 3).value + ">"
        if weapon_4star != "":
            weapon_4star = weapon_4star[:-1]
        else:
            weapon_4star = "无推荐"

        weapon_3star = ""
        for i in range(index, index + 5):
            if ws.cell(i, 4).value:
                weapon_3star += ws.cell(i, 4).value + ">"
        if weapon_3star != "":
            weapon_3star = weapon_3star[:-1]
        else:
            weapon_3star = "无推荐"

        artifacts = ""
        for i in range(index, index + 5):
            if ws.cell(i, 5).value:
                if ws.cell(i, 6).value:
                    artifacts += ws.cell(i, 5).value + "*2" + ws.cell(i, 6).value + "*2" + "\n"
                else:
                    artifacts += ws.cell(i, 5).value + "*4" + "\n"

        if artifacts != "":
            artifacts = artifacts[:-1]
        else:
            artifacts = "无推荐"

        im = char_adv_im.format(char_name, weapon_5star, weapon_4star, weapon_3star, artifacts)  # noqa
        return im
